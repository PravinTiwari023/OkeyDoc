import io
import os
import tempfile
import logging
import pandas as pd
import fitz
from openpyxl import load_workbook
from PIL import Image

logger = logging.getLogger(__name__)

def excel_to_pdf_service(excel_bytes, quality='high'):
    """
    Convert Excel document to PDF with specified quality settings.
    
    Args:
        excel_bytes: The input Excel file as bytes
        quality: Quality level ('low', 'medium', 'high')
        
    Returns:
        PDF file as bytes
    """
    try:
        # Create temporary files for input and output
        suffix = '.xlsx'
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temp_input:
            temp_input.write(excel_bytes)
            temp_input_path = temp_input.name

        # Create temporary output path
        temp_output_path = os.path.splitext(temp_input_path)[0] + '.pdf'

        try:
            # Load workbook using openpyxl
            workbook = load_workbook(temp_input_path)
            
            # Convert each worksheet to image
            images = []
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Save worksheet as temporary image
                img_path = f"{temp_input_path}_sheet_{len(images)}.png"
                
                # TODO: Implement worksheet to image conversion
                # For now, create a blank image as placeholder
                img = Image.new('RGB', (800, 600), 'white')
                img.save(img_path, 'PNG')
                images.append(img_path)

            # Create PDF from images
            images_pil = [Image.open(img_path) for img_path in images]
            if images_pil:
                images_pil[0].save(
                    temp_output_path,
                    'PDF',
                    save_all=True,
                    append_images=images_pil[1:]
                )

            # Read the output PDF
            with open(temp_output_path, 'rb') as pdf_file:
                pdf_bytes = pdf_file.read()

            return pdf_bytes

        finally:
            # Clean up temporary files
            try:
                os.remove(temp_input_path)
                if os.path.exists(temp_output_path):
                    os.remove(temp_output_path)
                for img_path in images:
                    if os.path.exists(img_path):
                        os.remove(img_path)
            except Exception as cleanup_error:
                logger.warning(f"Error cleaning up temporary files: {cleanup_error}")

    except Exception as e:
        logger.error(f"Excel to PDF conversion failed: {str(e)}")
        raise Exception(f"Excel to PDF conversion failed: {str(e)}")

def pdf_to_excel_service(pdf_bytes, quality='high'):
    """
    Convert PDF to Excel with specified quality
    
    Args:
        pdf_bytes: The input PDF file as bytes
        quality: Quality level ('low', 'medium', 'high')
        
    Returns:
        Excel file as bytes
    """
    try:
        # Create temporary files
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_pdf:
            temp_pdf.write(pdf_bytes)
            temp_pdf_path = temp_pdf.name

        temp_excel_path = temp_pdf_path.rsplit('.', 1)[0] + '.xlsx'

        try:
            # Open PDF with PyMuPDF
            pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
            
            # Extract tables from all pages
            all_tables = []
            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]
                try:
                    # Use find_tables() to detect and extract tables
                    tables = page.find_tables()
                    if tables:
                        for table in tables:
                            # Convert table to list of lists
                            table_data = table.extract()  # extract() returns the table as a list of lists
                            all_tables.append(pd.DataFrame(table_data))
                except AttributeError:
                    # Fallback if find_tables() isn't available (older PyMuPDF version)
                    logger.warning(f"Table extraction not supported in this PyMuPDF version for page {page_num}. Falling back to text extraction.")
                    text = page.get_text("text")
                    rows = [row.split('\t') for row in text.split('\n') if row.strip()]
                    if rows:
                        all_tables.append(pd.DataFrame(rows))

            # If no tables found, try text extraction as a last resort
            if not all_tables:
                for page_num in range(len(pdf_document)):
                    page = pdf_document[page_num]
                    text = page.get_text("text")
                    rows = [row.split('\t') for row in text.split('\n') if row.strip()]
                    if rows:
                        all_tables.append(pd.DataFrame(rows))

            if not all_tables:
                raise Exception("No tables or structured content found in PDF")

            # Quality settings affect how we process the data
            quality_settings = {
                'low': {'clean_data': False},
                'medium': {'clean_data': True},
                'high': {'clean_data': True}
            }
            settings = quality_settings.get(quality, quality_settings['high'])

            # Create Excel writer
            with pd.ExcelWriter(temp_excel_path, engine='openpyxl') as writer:
                for i, df in enumerate(all_tables):
                    if settings['clean_data']:
                        # Clean up data (remove empty rows/columns)
                        df = df.dropna(how='all').dropna(axis=1, how='all')
                    df.to_excel(writer, sheet_name=f'Sheet{i+1}', index=False, header=False)

            # Read the Excel file
            with open(temp_excel_path, 'rb') as excel_file:
                excel_bytes = excel_file.read()

            return excel_bytes

        finally:
            # Clean up temporary files
            try:
                os.remove(temp_pdf_path)
                if os.path.exists(temp_excel_path):
                    os.remove(temp_excel_path)
            except Exception as cleanup_error:
                logger.warning(f"Error cleaning up temporary files: {cleanup_error}")
            pdf_document.close()

    except Exception as e:
        logger.error(f"PDF to Excel conversion failed: {str(e)}")
        raise Exception(f"PDF to Excel conversion failed: {str(e)}")